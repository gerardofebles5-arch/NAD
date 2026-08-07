"""
Exportador de Resultados OCR
=============================
Exporta resultados OCR a múltiples formatos.

Funcionalidades:
  - Exportación a CSV
  - Exportación a Excel
  - Exportación a JSON
  - Exportación a PDF
"""

import csv
import json
from typing import Dict, List, Optional
from dataclasses import asdict
import os


class OCRResultExporter:
    """
    Exportador de resultados OCR a múltiples formatos.
    """
    
    def __init__(self):
        pass
    
    def export_to_csv(self, data: Dict, output_path: str):
        """
        Exporta resultados a CSV.
        
        Args:
            data: Datos de InvoiceData o diccionario
            output_path: Ruta del archivo CSV de salida
        """
        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Convertir data a diccionario si es InvoiceData
                if hasattr(data, 'to_dict'):
                    data = data.to_dict()
                
                # Escribir encabezados y valores
                for key, value in data.items():
                    writer.writerow([key, str(value)])
            
            print(f"[OK] Exportado a CSV: {output_path}")
        except Exception as e:
            print(f"[ERROR] Error exportando a CSV: {e}")
    
    def export_to_csv_table(self, data_list: List[Dict], output_path: str):
        """
        Exporta múltiples resultados a CSV en formato tabla.
        
        Args:
            data_list: Lista de diccionarios de datos
            output_path: Ruta del archivo CSV de salida
        """
        try:
            if not data_list:
                print("[WARN] No hay datos para exportar")
                return
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=data_list[0].keys())
                writer.writeheader()
                writer.writerows(data_list)
            
            print(f"[OK] Exportado a CSV tabla: {output_path}")
        except Exception as e:
            print(f"[ERROR] Error exportando a CSV tabla: {e}")
    
    def export_to_excel(self, data: Dict, output_path: str):
        """
        Exporta resultados a Excel (requiere openpyxl).
        
        Args:
            data: Datos de InvoiceData o diccionario
            output_path: Ruta del archivo Excel de salida
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultado OCR"
            
            # Convertir data a diccionario si es InvoiceData
            if hasattr(data, 'to_dict'):
                data = data.to_dict()
            
            # Escribir datos
            for row_idx, (key, value) in enumerate(data.items(), start=1):
                ws.cell(row=row_idx, column=1, value=key)
                ws.cell(row=row_idx, column=2, value=str(value))
            
            wb.save(output_path)
            print(f"[OK] Exportado a Excel: {output_path}")
        except ImportError:
            print("[WARN] openpyxl no instalado. Instala con: pip install openpyxl")
        except Exception as e:
            print(f"[ERROR] Error exportando a Excel: {e}")
    
    def export_to_excel_table(self, data_list: List[Dict], output_path: str):
        """
        Exporta múltiples resultados a Excel en formato tabla.
        
        Args:
            data_list: Lista de diccionarios de datos
            output_path: Ruta del archivo Excel de salida
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados OCR"
            
            if not data_list:
                print("[WARN] No hay datos para exportar")
                return
            
            # Escribir encabezados
            headers = list(data_list[0].keys())
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # Escribir datos
            for row_idx, data in enumerate(data_list, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = data.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            
            wb.save(output_path)
            print(f"[OK] Exportado a Excel tabla: {output_path}")
        except ImportError:
            print("[WARN] openpyxl no instalado. Instala con: pip install openpyxl")
        except Exception as e:
            print(f"[ERROR] Error exportando a Excel tabla: {e}")
    
    def export_to_json(self, data: Dict, output_path: str, indent: int = 2):
        """
        Exporta resultados a JSON.
        
        Args:
            data: Datos de InvoiceData o diccionario
            output_path: Ruta del archivo JSON de salida
            indent: Nivel de indentación
        """
        try:
            # Convertir data a diccionario si es InvoiceData
            if hasattr(data, 'to_dict'):
                data = data.to_dict()
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=indent, ensure_ascii=False)
            
            print(f"[OK] Exportado a JSON: {output_path}")
        except Exception as e:
            print(f"[ERROR] Error exportando a JSON: {e}")
    
    def export_to_pdf(self, data: Dict, output_path: str):
        """
        Exporta resultados a PDF (requiere reportlab).
        
        Args:
            data: Datos de InvoiceData o diccionario
            output_path: Ruta del archivo PDF de salida
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            
            # Convertir data a diccionario si es InvoiceData
            if hasattr(data, 'to_dict'):
                data = data.to_dict()
            
            doc = SimpleDocTemplate(output_path, pagesize=letter)
            elements = []
            
            # Crear tabla con datos
            table_data = [[key, str(value)] for key, value in data.items()]
            table = Table(table_data)
            
            # Estilo de tabla
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            print(f"[OK] Exportado a PDF: {output_path}")
        except ImportError:
            print("[WARN] reportlab no instalado. Instala con: pip install reportlab")
        except Exception as e:
            print(f"[ERROR] Error exportando a PDF: {e}")
    
    def export_all_formats(self, data: Dict, base_path: str):
        """
        Exporta resultados a todos los formatos disponibles.
        
        Args:
            data: Datos de InvoiceData o diccionario
            base_path: Ruta base sin extensión
        """
        # Crear directorio si no existe
        os.makedirs(os.path.dirname(base_path), exist_ok=True)
        
        # Exportar a cada formato
        self.export_to_csv(data, f"{base_path}.csv")
        self.export_to_json(data, f"{base_path}.json")
        
        try:
            self.export_to_excel(data, f"{base_path}.xlsx")
        except:
            pass
        
        try:
            self.export_to_pdf(data, f"{base_path}.pdf")
        except:
            pass


def export_ocr_result(data: Dict, output_path: str, format: str = 'json'):
    """
    Función de conveniencia para exportar resultados OCR.
    
    Args:
        data: Datos de InvoiceData o diccionario
        output_path: Ruta del archivo de salida
        format: Formato de exportación (csv, excel, json, pdf)
    """
    exporter = OCRResultExporter()
    
    if format == 'csv':
        exporter.export_to_csv(data, output_path)
    elif format == 'excel':
        exporter.export_to_excel(data, output_path)
    elif format == 'json':
        exporter.export_to_json(data, output_path)
    elif format == 'pdf':
        exporter.export_to_pdf(data, output_path)
    else:
        print(f"[ERROR] Formato no soportado: {format}")
