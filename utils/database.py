"""
Base de datos local de facturas — Libro de Compras/Ventas automático
======================================================================
Ni PhotoScan, ni CamScanner, ni MinerU llevan memoria de lo que escanean:
cada imagen es un evento aislado. Este módulo convierte NAD Scanner en una
herramienta de contabilidad real: cada factura procesada queda guardada,
buscable por RIF/fecha/número, y exportable a Excel/CSV para el cierre
mensual — que es exactamente el flujo que describe tu propio material de
onboarding ("Tu Negocio, Al Día").

Usa SQLite (stdlib, sin dependencias nuevas, un solo archivo portátil).
Diseño deliberadamente simple: una tabla, columnas planas, fácil de
inspeccionar con cualquier herramienta SQLite si algo falla.
"""

import os
import io
import json
import sqlite3
import base64
import threading
from datetime import datetime
from typing import Any, Dict, List, Optional
from contextlib import contextmanager

import cv2
import numpy as np

from utils.config import CONFIG

DB_PATH = os.path.join(CONFIG.output_dir, "nadscanner.db")

_lock = threading.RLock()  # RLock, no Lock: varias funciones de este módulo
# se llaman entre sí mientras ya sostienen el lock (ej. list_tenants() ->
# get_tenant_usage_summary() -> _connect() de nuevo). Con un Lock normal
# (no reentrante) eso es un deadlock garantizado del mismo hilo esperando
# a sí mismo — se confirmó reproduciendo el cuelgue real en /api/admin/tenants.

# Columnas que vienen directo de InvoiceData.to_dict() y se guardan tal cual
_INVOICE_FIELDS = [
    "numero_factura", "numero_control", "fecha", "rif_emisor", "razon_social",
    "direccion", "telefono", "base_imponible", "iva", "total", "condicion_pago",
    "cliente", "currency", "exchange_rate", "total_bs", "total_usd", "total_eur",
    "total_cop", "total_ars",
]

_SCHEMA = f"""
CREATE TABLE IF NOT EXISTS invoices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at TEXT NOT NULL,
    source TEXT NOT NULL DEFAULT 'guided',
    {", ".join(f"{f} TEXT" for f in _INVOICE_FIELDS)},
    ocr_confidence REAL DEFAULT 0,
    validation_errors TEXT DEFAULT '[]',
    qr_data TEXT DEFAULT NULL,
    thumbnail_b64 TEXT DEFAULT NULL,
    raw_text TEXT DEFAULT NULL,
    filename TEXT DEFAULT NULL
);
CREATE INDEX IF NOT EXISTS idx_invoices_rif ON invoices(rif_emisor);
CREATE INDEX IF NOT EXISTS idx_invoices_fecha ON invoices(fecha);
CREATE INDEX IF NOT EXISTS idx_invoices_numero ON invoices(numero_factura);
CREATE INDEX IF NOT EXISTS idx_invoices_created ON invoices(created_at);
"""


@contextmanager
def _connect():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    with _lock:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
            conn.commit()
        finally:
            conn.close()


def init_db():
    """Crea la tabla si no existe. Seguro llamar varias veces."""
    with _connect() as conn:
        conn.executescript(_SCHEMA)


def _make_thumbnail(image_b64: Optional[str], max_width: int = 220) -> Optional[str]:
    """
    Genera una miniatura JPEG pequeña en base64 a partir de la imagen
    realzada, para mostrar en el historial sin guardar la imagen completa
    (mantiene la base de datos liviana y evita duplicar almacenamiento
    de imágenes de alta resolución que ya existen en /tmp o en Drive).
    """
    if not image_b64:
        return None
    try:
        raw = base64.b64decode(image_b64)
        arr = np.frombuffer(raw, np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return None
        h, w = img.shape[:2]
        if w > max_width:
            scale = max_width / w
            img = cv2.resize(img, (max_width, int(h * scale)))
        ok, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 60])
        if not ok:
            return None
        return base64.b64encode(buf).decode("utf-8")
    except Exception:
        return None


def save_invoice(
    ocr_data: Dict[str, Any],
    validation_errors: Optional[List[str]] = None,
    ocr_confidence: float = 0.0,
    qr_data: Optional[Dict[str, Any]] = None,
    source: str = "guided",
    enhanced_image_b64: Optional[str] = None,
    filename: Optional[str] = None,
) -> int:
    """
    Guarda una factura procesada en el historial.

    Args:
        ocr_data: dict con las claves de InvoiceData.to_dict() (puede traer
                  claves extra, como 'raw_text' — se filtran automáticamente).
        validation_errors: lista de mensajes de validación (RIF, IVA, etc).
        ocr_confidence: confianza global del OCR (0-1).
        qr_data: datos extraídos de un código QR/barra, si se detectó uno.
        source: 'guided' (flujo de 5 tomas), 'batch' (lote), 'manual'.
        enhanced_image_b64: imagen realzada en base64 JPEG (se guarda solo
                             una miniatura, no la imagen completa).
        filename: nombre de archivo original, si aplica (modo lote).

    Returns:
        id de la fila insertada.
    """
    init_db()
    row = {f: ocr_data.get(f, "") for f in _INVOICE_FIELDS}
    for k in ("exchange_rate", "total_bs", "total_usd", "total_eur", "total_cop", "total_ars"):
        row[k] = str(row[k]) if row[k] not in (None, "") else "0"

    thumbnail = _make_thumbnail(enhanced_image_b64)

    with _connect() as conn:
        cols = ["created_at", "source"] + _INVOICE_FIELDS + [
            "ocr_confidence", "validation_errors", "qr_data", "thumbnail_b64",
            "raw_text", "filename",
        ]
        placeholders = ", ".join("?" for _ in cols)
        values = [
            datetime.now().isoformat(), source,
            *[row[f] for f in _INVOICE_FIELDS],
            float(ocr_confidence or 0),
            json.dumps(validation_errors or [], ensure_ascii=False),
            json.dumps(qr_data, ensure_ascii=False) if qr_data else None,
            thumbnail,
            (ocr_data.get("raw_text") or "")[:4000] or None,
            filename,
        ]
        cur = conn.execute(
            f"INSERT INTO invoices ({', '.join(cols)}) VALUES ({placeholders})",
            values,
        )
        return cur.lastrowid


def get_invoice(invoice_id: int) -> Optional[Dict[str, Any]]:
    init_db()
    with _connect() as conn:
        row = conn.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
        if not row:
            return None
        d = dict(row)
        d["validation_errors"] = json.loads(d.get("validation_errors") or "[]")
        d["qr_data"] = json.loads(d["qr_data"]) if d.get("qr_data") else None
        return d


def delete_invoice(invoice_id: int) -> bool:
    init_db()
    with _connect() as conn:
        cur = conn.execute("DELETE FROM invoices WHERE id = ?", (invoice_id,))
        return cur.rowcount > 0


def list_invoices(
    limit: int = 50,
    offset: int = 0,
    search: Optional[str] = None,
    rif: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    order: str = "created_at DESC",
) -> Dict[str, Any]:
    """
    Lista/busca facturas guardadas.

    Args:
        search: busca en razon_social, numero_factura, rif_emisor (LIKE).
        rif: filtra por RIF exacto.
        date_from/date_to: filtran por 'fecha' (string tal como quedó en
                            la factura — filtro simple, no parsea formatos).
    Returns:
        {"items": [...], "total": N}
    """
    init_db()
    limit = max(1, min(limit, 500))
    where = []
    params: List[Any] = []

    if search:
        where.append("(razon_social LIKE ? OR numero_factura LIKE ? OR rif_emisor LIKE ? OR cliente LIKE ?)")
        like = f"%{search}%"
        params += [like, like, like, like]
    if rif:
        where.append("rif_emisor = ?")
        params.append(rif)
    if date_from:
        where.append("fecha >= ?")
        params.append(date_from)
    if date_to:
        where.append("fecha <= ?")
        params.append(date_to)

    where_sql = f"WHERE {' AND '.join(where)}" if where else ""

    # Whitelist estricta para 'order' — nunca interpolar input del usuario
    # directamente en SQL (evita inyección vía el parámetro de orden).
    allowed_orders = {
        "created_at DESC", "created_at ASC",
        "fecha DESC", "fecha ASC",
        "total DESC", "total ASC",
    }
    if order not in allowed_orders:
        order = "created_at DESC"

    with _connect() as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM invoices {where_sql}", params).fetchone()[0]
        rows = conn.execute(
            f"SELECT * FROM invoices {where_sql} ORDER BY {order} LIMIT ? OFFSET ?",
            params + [limit, offset],
        ).fetchall()
        items = []
        for r in rows:
            d = dict(r)
            d["validation_errors"] = json.loads(d.get("validation_errors") or "[]")
            d["qr_data"] = json.loads(d["qr_data"]) if d.get("qr_data") else None
            items.append(d)
        return {"items": items, "total": total, "limit": limit, "offset": offset}


def get_summary(date_from: Optional[str] = None, date_to: Optional[str] = None) -> Dict[str, Any]:
    """
    Resumen contable rápido: cantidad de facturas, totales por moneda,
    RIFs más frecuentes. Pensado para un mini-dashboard del cierre mensual.
    """
    init_db()
    where = []
    params: List[Any] = []
    if date_from:
        where.append("fecha >= ?")
        params.append(date_from)
    if date_to:
        where.append("fecha <= ?")
        params.append(date_to)
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""

    with _connect() as conn:
        count = conn.execute(f"SELECT COUNT(*) FROM invoices {where_sql}", params).fetchone()[0]

        by_currency = {}
        rows = conn.execute(
            f"SELECT currency, COUNT(*) as n, "
            f"SUM(CAST(REPLACE(REPLACE(total,'.',''),',','.') AS REAL)) as suma "
            f"FROM invoices {where_sql} GROUP BY currency",
            params,
        ).fetchall()
        for r in rows:
            curr = r["currency"] or "N/D"
            by_currency[curr] = {
                "count": r["n"],
                "total": round(r["suma"], 2) if r["suma"] is not None else 0,
            }

        top_rifs = conn.execute(
            f"SELECT rif_emisor, razon_social, COUNT(*) as n FROM invoices "
            f"{where_sql} {'AND' if where else 'WHERE'} rif_emisor != '' "
            f"GROUP BY rif_emisor ORDER BY n DESC LIMIT 10",
            params,
        ).fetchall()

        avg_confidence = conn.execute(
            f"SELECT AVG(ocr_confidence) FROM invoices {where_sql}", params
        ).fetchone()[0]

        return {
            "total_facturas": count,
            "por_moneda": by_currency,
            "top_emisores": [dict(r) for r in top_rifs],
            "confianza_promedio": round(avg_confidence, 4) if avg_confidence else 0,
        }


def export_to_excel(
    search: Optional[str] = None,
    rif: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> bytes:
    """
    Genera un libro de Excel con TODAS las facturas que matcheen el filtro
    (sin paginar — es para exportar el mes completo de una vez).

    Returns:
        Bytes del archivo .xlsx, listo para servir como descarga.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    result = list_invoices(limit=100000, search=search, rif=rif, date_from=date_from, date_to=date_to)
    items = result["items"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = [
        "Fecha", "N° Factura", "N° Control", "RIF Emisor", "Razón Social",
        "Cliente", "Base Imponible", "IVA", "Total", "Moneda",
        "Tasa de cambio", "Condición de Pago", "Confianza OCR",
        "Advertencias", "Registrado el",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="D9B158", end_color="D9B158", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="241609")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for inv in items:
        ws.append([
            inv.get("fecha", ""),
            inv.get("numero_factura", ""),
            inv.get("numero_control", ""),
            inv.get("rif_emisor", ""),
            inv.get("razon_social", ""),
            inv.get("cliente", ""),
            inv.get("base_imponible", ""),
            inv.get("iva", ""),
            inv.get("total", ""),
            inv.get("currency", ""),
            inv.get("exchange_rate", ""),
            inv.get("condicion_pago", ""),
            f"{(inv.get('ocr_confidence') or 0) * 100:.0f}%",
            "; ".join(inv.get("validation_errors") or []),
            (inv.get("created_at") or "")[:19].replace("T", " "),
        ])

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(45, max_len + 3)

    summary = get_summary(date_from=date_from, date_to=date_to)
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Total de facturas", summary["total_facturas"]])
    ws2.append(["Confianza OCR promedio", f"{summary['confianza_promedio']*100:.1f}%"])
    ws2.append([])
    ws2.append(["Moneda", "Cantidad", "Total"])
    for curr, data in summary["por_moneda"].items():
        ws2.append([curr, data["count"], data["total"]])
    ws2.append([])
    ws2.append(["RIF más frecuente", "Razón social", "Cantidad"])
    for r in summary["top_emisores"]:
        ws2.append([r["rif_emisor"], r["razon_social"], r["n"]])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_csv(
    search: Optional[str] = None,
    rif: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> str:
    """Genera un CSV plano (para importar en cualquier sistema contable)."""
    import csv

    result = list_invoices(limit=100000, search=search, rif=rif, date_from=date_from, date_to=date_to)
    items = result["items"]

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "fecha", "numero_factura", "numero_control", "rif_emisor", "razon_social",
        "cliente", "base_imponible", "iva", "total", "currency", "exchange_rate",
        "condicion_pago", "ocr_confidence", "created_at",
    ])
    for inv in items:
        writer.writerow([
            inv.get("fecha", ""), inv.get("numero_factura", ""), inv.get("numero_control", ""),
            inv.get("rif_emisor", ""), inv.get("razon_social", ""), inv.get("cliente", ""),
            inv.get("base_imponible", ""), inv.get("iva", ""), inv.get("total", ""),
            inv.get("currency", ""), inv.get("exchange_rate", ""), inv.get("condicion_pago", ""),
            inv.get("ocr_confidence", ""), inv.get("created_at", ""),
        ])
    return buf.getvalue()
